from shelve import Shelf
from sre_constants import SUCCESS
from weakref import finalize
import dash
from dash.dependencies import Output, Input
from numpy import true_divide
from openpyxl import load_workbook
import dash_core_components as dcc
import dash_html_components as html
from dashboard.checkdate import is_date
from dashboard.datetodate import datetodate
from dash.exceptions import PreventUpdate
import time
import serial

#e7e7e7 is gray base

import_button_style = {'height':'40px', 'width':'150px', 'background-color':'#e7e7e7'}
export_button_style = {'height':'40px', 'width':'150px', 'background-color':'#e7e7e7'}
submit_button_style = {'width':'200px','margin-left':'120px', 'background-color':'#e7e7e7'}
finalize_button_style = {'width':'200px','margin-left':'120px', 'background-color':'#e7e7e7'}

n = 0
if_import = False
if_export = False
Shelf_Colors = {}
samples = 0
sample_location = []

wb = load_workbook('/Users/ortho/OneDrive - stevens.edu/Documents/GitHub/SD_Infineum/Senior Design Dashboard/Shelf Data.xlsx')
ws = wb['Shelf 1']

## updates what is currently selected

def register(app:dash):

    @app.callback(
        Output('Current_shelf', 'children'),
        [Input('shelves', 'value'),
        Input('finalize_button', 'n_clicks')]
        )
    def DisplayCurrentShelf(current_shelf, finalization):

        if current_shelf == 'Shelf_1':
            for i in range(2,47):
                if ws.cell(row = i, column = 2).value is None:
                    Shelf_Colors['Slot_' + str(i-1)+ '_Color'] = 'white'
                else:
                    Shelf_Colors['Slot_' + str(i-1) + '_Color'] = 'black'
            return html.Div([
                    html.Div([
                        html.Label('Shelf 1'),
                    ], style={'height': '40px','width':'100px', 'font-size':'30px'}),

                    html.Div([
                        html.Div([
                            html.Button(id='Slot_1', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color': Shelf_Colors['Slot_1_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_2', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_2_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_3', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_3_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_4', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_4_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_5', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_5_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_6', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_6_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                        html.Button(id='Slot_7', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_7_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                        html.Button(id='Slot_8', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_8_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                        html.Button(id='Slot_9', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_9_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                    ], style={'height':'70px'}),


                    html.Div([
                        html.Div([
                            html.Button(id='Slot_10', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_10_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_11', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_11_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_12', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_12_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_13', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_13_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_14', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_14_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_15', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_15_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                        html.Button(id='Slot_16', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_16_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                        html.Button(id='Slot_17', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_17_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                        html.Button(id='Slot_18', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_18_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                    ], style={'height':'70px'}),


                    html.Div([
                        html.Div([
                            html.Button(id='Slot_19', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_19_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_20', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_20_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_21', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_21_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_22', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_22_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_23', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_23_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_24', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_24_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                        html.Button(id='Slot_25', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_25_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                        html.Button(id='Slot_26', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_26_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                        html.Button(id='Slot_27', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_27_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                    ], style={'height':'70px'}),


                    html.Div([
                        html.Div([
                            html.Button(id='Slot_28', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_28_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_29', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_29_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_30', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_30_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_31', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_31_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_32', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_32_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_33', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_33_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                        html.Button(id='Slot_34', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_34_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                        html.Button(id='Slot_35', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_35_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                        html.Button(id='Slot_36', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_36_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                    ], style={'height':'70px'}),


                    html.Div([
                        html.Div([
                            html.Button(id='Slot_37', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_37_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_38', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_38_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_39', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_39_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_40', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_40_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_41', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_41_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                            html.Button(id='Slot_42', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_42_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                        html.Button(id='Slot_43', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_43_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                        html.Button(id='Slot_44', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_44_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                        html.Div([
                        html.Button(id='Slot_45', n_clicks=0, style={'border-radius':'50%', 'height':'50px', 'width':'50px', 'border': '2px solid black', 'background-color':Shelf_Colors['Slot_45_Color']})
                        ], style = {'display':'inline-block', 'width':'70px'}),
                    ], style={'height':'70px'}),

                    html.Div([
                        html.Table([
                            html.Tr([
                                html.Th(str(ws.cell(row = 1, column = 2).value)),
                                html.Th(str(ws.cell(row = 1, column = 3).value)),
                                html.Th(str(ws.cell(row = 1, column = 4).value)),
                            ]),
                            html.Tr([
                                html.Th(id = 'sample_info_column_1', children=['No sample Selected']),
                                html.Th(id = 'sample_info_column_2', children=['']),
                                html.Th(id = 'sample_info_column_3', children=['']),
                            ]),
                        ], style={'Boarder':'solid', 'text-align':'center', 'font-size':'20px', 'margin-left': 'auto', 'margin-right' : 'auto'})
                    ], style={'vertical-align':'top'})
                ], style={})

    @app.callback([
            Output('sample_info_column_1', 'children'),
            Output('sample_info_column_2', 'children'),
            Output('sample_info_column_3', 'children')],
            [Input('Slot_1', 'n_clicks'),
            Input('Slot_2', 'n_clicks'),
            Input('Slot_3', 'n_clicks'),
            Input('Slot_4', 'n_clicks'),
            Input('Slot_5', 'n_clicks'),
            Input('Slot_6', 'n_clicks'),
            Input('Slot_7', 'n_clicks'),
            Input('Slot_8', 'n_clicks'),
            Input('Slot_9', 'n_clicks'),
            Input('Slot_10', 'n_clicks'),
            Input('Slot_11', 'n_clicks'),
            Input('Slot_12', 'n_clicks'),
            Input('Slot_13', 'n_clicks'),
            Input('Slot_14', 'n_clicks'),
            Input('Slot_15', 'n_clicks'),
            Input('Slot_16', 'n_clicks'),
            Input('Slot_17', 'n_clicks'),
            Input('Slot_18', 'n_clicks'),
            Input('Slot_19', 'n_clicks'),
            Input('Slot_20', 'n_clicks'),
            Input('Slot_21', 'n_clicks'),
            Input('Slot_22', 'n_clicks'),
            Input('Slot_23', 'n_clicks'),
            Input('Slot_24', 'n_clicks'),
            Input('Slot_25', 'n_clicks'),
            Input('Slot_26', 'n_clicks'),
            Input('Slot_27', 'n_clicks'),
            Input('Slot_28', 'n_clicks'),
            Input('Slot_29', 'n_clicks'),
            Input('Slot_30', 'n_clicks'),
            Input('Slot_31', 'n_clicks'),
            Input('Slot_32', 'n_clicks'),
            Input('Slot_33', 'n_clicks'),
            Input('Slot_34', 'n_clicks'),
            Input('Slot_35', 'n_clicks'),
            Input('Slot_36', 'n_clicks'),
            Input('Slot_37', 'n_clicks'),
            Input('Slot_38', 'n_clicks'),
            Input('Slot_39', 'n_clicks'),
            Input('Slot_40', 'n_clicks'),
            Input('Slot_41', 'n_clicks'),
            Input('Slot_42', 'n_clicks'),
            Input('Slot_43', 'n_clicks'),
            Input('Slot_44', 'n_clicks'),
            Input('Slot_45', 'n_clicks'),
            ])
    def Display_Shelves(
    Slot_1,Slot_2,Slot_3,Slot_4,Slot_5,Slot_6,Slot_7,Slot_8,Slot_9,Slot_10,
    Slot_11,Slot_12,Slot_13,Slot_14,Slot_15,Slot_16,Slot_17,Slot_18,Slot_19,Slot_20,
    Slot_21,Slot_22,Slot_23,Slot_24,Slot_25,Slot_26,Slot_27,Slot_28,Slot_29,Slot_30,
    Slot_31,Slot_32,Slot_33,Slot_34,Slot_35,Slot_36,Slot_37,Slot_38,Slot_39,Slot_40,
    Slot_41,Slot_42,Slot_43,Slot_44,Slot_45):
        global n, arduino
        ELN_number = 'No sample selected'
        Arrial_date = ''
        experation_date = ''

        if n == 0:
            arduino = serial.Serial(port='COM4', baudrate=115200, timeout=.1)
            n = n + 1
            raise PreventUpdate
            
        else:
            update_sample_info = dash.callback_context
            current_sample = update_sample_info.triggered[0]['prop_id'].split('.')[0]
            for i in range(2,47):
                if current_sample == 'Slot_' + str(i-1):
                    if Shelf_Colors['Slot_' + str(i-1) + '_Color'] == 'black':
                        ELN_number = str(ws.cell(row = i, column = 2).value)
                        Arrial_date = str(ws.cell(row = i, column = 3).value)[:-9]
                        experation_date = str(ws.cell(row = i, column = 4).value)[:-9]
                    else:
                        ELN_number = 'No sample found'
                        Arrial_date = ''
                        experation_date =''

            return [ELN_number, Arrial_date, experation_date]




    @app.callback(
        [Output('import','style'),
        Output('export','style'),
        Output('user_entry', 'children')],
        [Input('submit_button','n_clicks'),
        Input('import','n_clicks'),
        Input('export','n_clicks'),
        Input('finalize_button', 'n_clicks'),
        Input('input_1_placeholder', 'value'),
        Input('input_1_ELN', 'value'),
        Input('input_1_arrival', 'value'),
        Input('input_1_experation', 'value'),
        Input('input_2_placeholder', 'value'),
        Input('input_2_ELN', 'value'),
        Input('input_2_arrival', 'value'),
        Input('input_2_experation', 'value'),
        Input('input_3_placeholder', 'value'),
        Input('input_3_ELN', 'value'),
        Input('input_3_arrival', 'value'),
        Input('input_3_experation', 'value'),
        Input('input_4_placeholder', 'value'),
        Input('input_4_ELN', 'value'),
        Input('input_4_arrival', 'value'),
        Input('input_4_experation', 'value')])
    def submission(submit_button, importded, exported, fialize_button,
        input_1_placeholder, input_1_ELN, input_1_arrival, input_1_experation,
        input_2_placeholder, input_2_ELN, input_2_arrival, input_2_experation,
        input_3_placeholder, input_3_ELN, input_3_arrival, input_3_experation,
        input_4_placeholder, input_4_ELN, input_4_arrival, input_4_experation):

        global if_export, if_import, import_button_style, export_button_style, samples, sample_location

        submission = dash.callback_context
        submission_status = submission.triggered[0]['prop_id'].split('.')[0]

        date_check_arrival = False
        date_check_experation = False
        user_entry = dash.no_update
        if submission_status == 'import':
            if_export = False
            if_import = True
            print('Importing Now')
            import_button_style['background-color'] = 'gray'
            export_button_style['background-color'] = '#e7e7e7'

        elif submission_status == 'export':
            if_export = True
            if_import = False
            print('Exporting Now')
            import_button_style['background-color'] = '#e7e7e7'
            export_button_style['background-color'] = 'gray'

        elif submission_status == 'submit_button':

            if if_export == True or if_import == True:

                try:
                    input_1_placeholder = int(input_1_placeholder)
                except:
                    hello_world = 0
                try:
                    input_1_ELN = int(input_1_ELN)
                except:
                    hello_world = 0


                       
                if if_import == True:
                    if type(input_1_placeholder) == int and type(input_1_ELN) == int and is_date(input_1_arrival) == True and is_date(input_1_experation) == True:

                        time.sleep(2)
                        arduino.write(bytes('LED', 'utf-8'))
                        time.sleep(2)
                        arduino.write(bytes(str(input_1_placeholder), 'utf-8'))

                        if ws.cell(row = input_1_placeholder+1, column= 2).value == None:
                            ws.cell(row = input_1_placeholder+1, column= 2).value = input_1_ELN
                            ws.cell(row = input_1_placeholder+1, column= 3).value = datetodate(input_1_arrival)
                            ws.cell(row = input_1_placeholder+1, column= 4).value = datetodate(input_1_experation)
                        
                        samples = samples + 1

                        sample_location.append(input_1_placeholder)
                    
                    if type(input_2_placeholder) == int and type(input_2_ELN) == int and is_date(input_2_arrival) == True and is_date(input_2_experation) == True:

                        time.sleep(2)
                        arduino.write(bytes('LED', 'utf-8'))
                        time.sleep(2)
                        arduino.write(bytes(str(input_2_placeholder), 'utf-8'))

                        if ws.cell(row = input_2_placeholder+1, column= 2).value == None:
                            ws.cell(row = input_2_placeholder+1, column= 2).value = input_1_ELN
                            ws.cell(row = input_2_placeholder+1, column= 3).value = datetodate(input_2_arrival)
                            ws.cell(row = input_2_placeholder+1, column= 4).value = datetodate(input_2_experation)

                        samples = samples + 1

                        sample_location.append(input_2_placeholder)

                    if type(input_3_placeholder) == int and type(input_3_ELN) == int and is_date(input_3_arrival) == True and is_date(input_3_experation) == True:

                        time.sleep(2)
                        arduino.write(bytes('LED', 'utf-8'))
                        time.sleep(2)
                        arduino.write(bytes(str(input_3_placeholder), 'utf-8'))

                        if ws.cell(row = input_3_placeholder+1, column= 2).value == None:
                            ws.cell(row = input_3_placeholder+1, column= 2).value = input_1_ELN
                            ws.cell(row = input_3_placeholder+1, column= 3).value = datetodate(input_3_arrival)
                            ws.cell(row = input_3_placeholder+1, column= 4).value = datetodate(input_3_experation)

                        samples = samples + 1

                        sample_location.append(input_3_placeholder)

                    if type(input_4_placeholder) == int and type(input_4_ELN) == int and is_date(input_4_arrival) == True and is_date(input_4_experation) == True:

                        time.sleep(2)
                        arduino.write(bytes('LED', 'utf-8'))
                        time.sleep(2)
                        arduino.write(bytes(str(input_4_placeholder), 'utf-8'))

                        if ws.cell(row = input_4_placeholder+1, column= 2).value == None:
                            ws.cell(row = input_4_placeholder+1, column= 2).value = input_1_ELN
                            ws.cell(row = input_4_placeholder+1, column= 3).value = datetodate(input_4_arrival)
                            ws.cell(row = input_4_placeholder+1, column= 4).value = datetodate(input_4_experation)
                    
                        samples = samples + 1

                        sample_location.append(input_4_placeholder)
                    
                    if samples > 0:
                        user_entry = html.Div(
                            [html.Label('Please place the samples into the lit up spots')],
                            style = {'margin-left':'120px'})  
                        import_button_style['background-color'] = '#e7e7e7'
                        export_button_style['background-color'] = '#e7e7e7'
                        if_import = False
                        wb.save('C:/Users/ortho/OneDrive - stevens.edu/Documents/GitHub/SD_Infineum/Senior Design Dashboard/Shelf Data.xlsx')


                if if_export == True:
                    if type(input_1_placeholder) == int and type(input_1_ELN) == int and is_date(input_1_arrival) == True and is_date(input_1_experation) == True:
                        if  str(ws.cell(row = input_1_placeholder + 1, column = 2).value) == str(input_1_ELN) and str(ws.cell(row = input_1_placeholder + 1, column = 3).value) == datetodate(input_1_arrival) and str(ws.cell(row = input_1_placeholder + 1, column = 4).value) == datetodate(input_1_experation):
                            time.sleep(2)
                            arduino.write(bytes('LED', 'utf-8'))
                            time.sleep(2)
                            arduino.write(bytes(str(input_1_placeholder), 'utf-8'))
                    
                            ws.cell(row = input_1_placeholder + 1, column= 2).value = None
                            ws.cell(row = input_1_placeholder + 1, column= 3).value = None
                            ws.cell(row = input_1_placeholder + 1, column= 4).value = None
                        samples = samples + 1
                        sample_location.append(input_1_placeholder)

                    if type(input_2_placeholder) == int and type(input_2_ELN) == int and is_date(input_2_arrival) == True and is_date(input_2_experation) == True:
                        if  str(ws.cell(row = input_2_placeholder + 1, column = 2).value) == str(input_2_ELN) and str(ws.cell(row = input_2_placeholder + 1, column = 3).value) == datetodate(input_2_arrival) and str(ws.cell(row = input_2_placeholder + 1, column = 4).value) == datetodate(input_2_experation):
                            time.sleep(2)
                            arduino.write(bytes('LED', 'utf-8'))
                            time.sleep(2)
                            arduino.write(bytes(str(input_2_placeholder), 'utf-8'))
                            ws.cell(row = input_2_placeholder + 1, column= 2).value = None
                            ws.cell(row = input_2_placeholder + 1, column= 3).value = None
                            ws.cell(row = input_2_placeholder + 1, column= 4).value = None
                        samples = samples + 1
                        sample_location.append(input_2_placeholder)

                    if type(input_3_placeholder) == int and type(input_3_ELN) == int and is_date(input_3_arrival) == True and is_date(input_3_experation) == True:
                        if  str(ws.cell(row = input_3_placeholder + 1, column = 2).value) == str(input_3_ELN) and str(ws.cell(row = input_3_placeholder + 1, column = 3).value) == datetodate(input_3_arrival) and str(ws.cell(row = input_3_placeholder + 1, column = 4).value) == datetodate(input_3_experation):
                            time.sleep(2)
                            arduino.write(bytes('LED', 'utf-8'))
                            time.sleep(2)
                            arduino.write(bytes(str(input_3_placeholder), 'utf-8'))
                            ws.cell(row = input_3_placeholder + 1, column= 2).value = None
                            ws.cell(row = input_3_placeholder + 1, column= 3).value = None
                            ws.cell(row = input_3_placeholder + 1, column= 4).value = None
                        samples = samples + 1
                        sample_location.append(input_3_placeholder)
                    
                    if type(input_4_placeholder) == int and type(input_4_ELN) == int and is_date(input_4_arrival) == True and is_date(input_4_experation) == True:
                        if  str(ws.cell(row = input_4_placeholder + 1, column = 2).value) == str(input_4_ELN) and str(ws.cell(row = input_4_placeholder + 1, column = 3).value) == datetodate(input_4_arrival) and str(ws.cell(row = input_4_placeholder + 1, column = 4).value) == datetodate(input_4_experation):
                            time.sleep(2)
                            arduino.write(bytes('LED', 'utf-8'))
                            time.sleep(2)
                            arduino.write(bytes(str(input_4_placeholder), 'utf-8'))
                            ws.cell(row = input_4_placeholder + 1, column= 2).value = None
                            ws.cell(row = input_4_placeholder + 1, column= 3).value = None
                            ws.cell(row = input_4_placeholder + 1, column= 4).value = None
                        samples = samples + 1
                        sample_location.append(input_4_placeholder)
                    
                    if samples > 0:
                        user_entry = html.Div(
                            [html.Label('Please take all the samples that are lit up')],
                            style = {'margin-left':'120px'})  
                        if_export = False
                        import_button_style['background-color'] = '#e7e7e7'
                        export_button_style['background-color'] = '#e7e7e7'
                        wb.save('C:/Users/ortho/OneDrive - stevens.edu/Documents/GitHub/SD_Infineum/Senior Design Dashboard/Shelf Data.xlsx')

        if submission_status == 'finalize_button':
            
            time.sleep(2)
            arduino.write(bytes('LED', 'utf-8'))
            time.sleep(2)
            arduino.write(bytes(str(input_1_placeholder), 'utf-8'))
            
            if_export = False
            if_import = False
            samples = 0
            
            #if samples > 0:
            #    for i in range(0,samples):
            #        time.sleep(2)
            #        arduino.write(bytes('Sensor', 'utf-8'))
            #        time.sleep(2)
                    #arduino.write(bytes(sample_location[i], 'utf-8'))

        
        return [import_button_style, export_button_style, user_entry]