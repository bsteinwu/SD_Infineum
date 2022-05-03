from openpyxl import load_workbook
from checkdate import is_date
from datetodate import datetodate

input_1_placeholder = 1
input_1_ELN = 12345678
input_1_arrival = '1/15/22'
input_1_experation = '5/3/25'


wb = load_workbook('/Users/ortho/OneDrive - stevens.edu/Documents/GitHub/SD_Infineum/Senior Design Dashboard/Shelf Data.xlsx')
ws = wb['Shelf 1']

if type(input_1_placeholder) == int and type(input_1_ELN) == int and is_date(input_1_arrival) == True and is_date(input_1_experation) == True:
                    print('Check Correct')
                    print('exported samples')
                    print(ws.cell(row = input_1_placeholder + 1, column = 2).value)
                    print(str(input_1_ELN))
                    print(str(ws.cell(row = input_1_placeholder + 1, column = 3).value))
                    print(datetodate(input_1_arrival))
                    print(str(ws.cell(row = input_1_placeholder + 1, column = 4).value))
                    print(datetodate(input_1_experation))
                    
                    if str(ws.cell(row = input_1_placeholder + 1, column = 2).value) == str(input_1_ELN):
                        print('part 1 works')
                        if str(ws.cell(row = input_1_placeholder + 1, column = 3).value) == datetodate(input_1_arrival):
                            print('part 2 works')
                            if str(ws.cell(row = input_1_placeholder + 1, column = 4).value) == datetodate(input_1_experation):
                                print('part 3 works')
                                ws.cell(row = input_1_placeholder + 1, column= 2).value = None
                                ws.cell(row = input_1_placeholder + 1, column= 3).value = None
                                ws.cell(row = input_1_placeholder + 1, column= 4).value = None


print('finished exporting')
print(ws.cell(row = input_1_placeholder + 1, column= 2).value)
print(ws.cell(row = input_1_placeholder + 1, column= 3).value)
print(ws.cell(row = input_1_placeholder + 1, column= 4).value)

wb.save('C:/Users/ortho/OneDrive - stevens.edu/Documents/GitHub/SD_Infineum/Senior Design Dashboard/Shelf Data.xlsx')